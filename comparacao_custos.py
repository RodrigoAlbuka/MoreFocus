#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def create_cost_comparison_spreadsheet():
    # Criar workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    original_fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
    free_tier_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    savings_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    currency_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ABA 1: COMPARAÇÃO DETALHADA
    ws_comparison = wb.create_sheet("Comparação Detalhada")
    
    # Cabeçalho
    ws_comparison['A1'] = "COMPARAÇÃO DE CUSTOS - INFRAESTRUTURA ORIGINAL vs FREE TIER"
    ws_comparison['A1'].font = Font(bold=True, size=16)
    ws_comparison.merge_cells('A1:G1')
    
    # Headers
    headers = ['Categoria', 'Serviço', 'Original (USD/mês)', 'Free Tier (USD/mês)', 'Economia (USD)', 'Economia (%)', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws_comparison.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Dados de comparação
    comparison_data = [
        # COMPUTE RESOURCES
        ["Compute", "EKS Cluster Control Plane", 216, 0, 216, 100, "Substituído por Docker local"],
        ["Compute", "EC2 t3.large (6x)", 402, 0, 402, 100, "Free tier: 1x t3.micro"],
        ["Compute", "EC2 c5.xlarge (3x)", 438, 0, 438, 100, "Não necessário no free tier"],
        ["Compute", "Lambda Functions", 0, 0, 0, 0, "Incluído no free tier"],
        
        # DATABASE & STORAGE
        ["Database", "RDS PostgreSQL r5.xlarge (3x)", 855, 0, 855, 100, "Free tier: 1x t3.micro"],
        ["Database", "ElastiCache Redis (3x)", 426, 0, 426, 100, "Substituído por Redis local"],
        ["Storage", "S3 Storage (5TB)", 115, 0, 115, 100, "Free tier: 5GB incluídos"],
        ["Storage", "EFS Storage (1TB)", 300, 0, 300, 100, "Substituído por volumes locais"],
        
        # NETWORKING
        ["Network", "Application Load Balancer (3x)", 66, 0, 66, 100, "Substituído por Nginx"],
        ["Network", "CloudFront CDN", 85, 0, 85, 100, "Free tier: 1TB incluído"],
        ["Network", "Data Transfer", 10, 0, 10, 100, "Free tier: 15GB incluídos"],
        
        # MONITORING & SECURITY
        ["Monitor", "CloudWatch", 150, 0, 150, 100, "Substituído por soluções gratuitas"],
        ["Security", "AWS WAF", 15, 0, 15, 100, "Nginx rate limiting"],
        ["Security", "Secrets Manager", 20, 0, 20, 100, "Variáveis de ambiente"],
        
        # THIRD-PARTY SERVICES
        ["APM", "New Relic", 250, 0, 250, 100, "UptimeRobot gratuito"],
        ["Monitor", "DataDog", 225, 0, 225, 100, "Grafana Cloud gratuito"],
        ["CI/CD", "GitLab Premium", 99, 0, 99, 100, "GitHub Actions gratuito"],
        
        # CUSTOS OPCIONAIS FREE TIER
        ["Domain", "Domínio .com", 0, 1, -1, 0, "Opcional - $12/ano"],
        ["CDN", "Cloudflare Pro", 0, 20, -20, 0, "Opcional para performance"],
    ]
    
    total_original = 0
    total_free_tier = 0
    
    # Preencher dados
    for row, data in enumerate(comparison_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws_comparison.cell(row=row, column=col, value=value)
            cell.border = thin_border
            
            if col in [3, 4, 5]:  # Colunas de valores monetários
                cell.alignment = currency_alignment
                if col == 3:  # Original
                    total_original += value
                    if value > 0:
                        cell.fill = original_fill
                        cell.font = Font(color="FFFFFF", bold=True)
                elif col == 4:  # Free Tier
                    total_free_tier += value
                    if value == 0:
                        cell.fill = free_tier_fill
                        cell.font = Font(color="FFFFFF", bold=True)
                elif col == 5:  # Economia
                    if value > 0:
                        cell.fill = savings_fill
                        cell.font = Font(color="FFFFFF", bold=True)
            elif col == 6:  # Porcentagem
                cell.alignment = center_alignment
                if value == 100:
                    cell.fill = savings_fill
                    cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.alignment = center_alignment if col in [1, 6] else Alignment(horizontal="left", vertical="center")
    
    # Totais
    total_row = len(comparison_data) + 5
    ws_comparison.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=14)
    ws_comparison.cell(row=total_row, column=2, value="Mensal").font = Font(bold=True, size=14)
    
    cell_original = ws_comparison.cell(row=total_row, column=3, value=total_original)
    cell_original.font = Font(bold=True, size=14, color="FFFFFF")
    cell_original.fill = original_fill
    cell_original.alignment = currency_alignment
    
    cell_free = ws_comparison.cell(row=total_row, column=4, value=total_free_tier)
    cell_free.font = Font(bold=True, size=14, color="FFFFFF")
    cell_free.fill = free_tier_fill
    cell_free.alignment = currency_alignment
    
    total_savings = total_original - total_free_tier
    cell_savings = ws_comparison.cell(row=total_row, column=5, value=total_savings)
    cell_savings.font = Font(bold=True, size=14, color="FFFFFF")
    cell_savings.fill = savings_fill
    cell_savings.alignment = currency_alignment
    
    savings_percent = (total_savings / total_original * 100) if total_original > 0 else 0
    cell_percent = ws_comparison.cell(row=total_row, column=6, value=f"{savings_percent:.1f}%")
    cell_percent.font = Font(bold=True, size=14)
    cell_percent.alignment = center_alignment
    
    # Ajustar larguras
    comparison_widths = [12, 25, 18, 18, 15, 12, 30]
    for col, width in enumerate(comparison_widths, 1):
        ws_comparison.column_dimensions[get_column_letter(col)].width = width
    
    # ABA 2: ANÁLISE TEMPORAL
    ws_temporal = wb.create_sheet("Análise Temporal")
    
    # Cabeçalho
    ws_temporal['A1'] = "ANÁLISE DE CUSTOS AO LONGO DO TEMPO"
    ws_temporal['A1'].font = Font(bold=True, size=16)
    ws_temporal.merge_cells('A1:H1')
    
    # Headers temporais
    temporal_headers = ['Período', 'Original (USD)', 'Free Tier (USD)', 'Economia (USD)', 'Economia Acumulada', 'ROI (%)', 'Observações', 'Recomendação']
    for col, header in enumerate(temporal_headers, 1):
        cell = ws_temporal.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Dados temporais (considerando que free tier é válido por 12 meses)
    temporal_data = [
        ["Mês 1", 3672, 21, 3651, 3651, 17385, "Setup inicial", "Use Free Tier"],
        ["Mês 2", 3672, 21, 3651, 7302, 17385, "Desenvolvimento", "Use Free Tier"],
        ["Mês 3", 3672, 21, 3651, 10953, 17385, "MVP Completo", "Use Free Tier"],
        ["Mês 4", 3672, 21, 3651, 14604, 17385, "Primeiros clientes", "Use Free Tier"],
        ["Mês 5", 3672, 21, 3651, 18255, 17385, "Validação", "Use Free Tier"],
        ["Mês 6", 3672, 21, 3651, 21906, 17385, "Otimização", "Use Free Tier"],
        ["Mês 7", 3672, 21, 3651, 25557, 17385, "Expansão Argentina", "Use Free Tier"],
        ["Mês 8", 3672, 21, 3651, 29208, 17385, "Consolidação", "Use Free Tier"],
        ["Mês 9", 3672, 21, 3651, 32859, 17385, "Lançamento EUA", "Use Free Tier"],
        ["Mês 10", 3672, 21, 3651, 36510, 17385, "Crescimento", "Use Free Tier"],
        ["Mês 11", 3672, 21, 3651, 40161, 17385, "Preparação Europa", "Use Free Tier"],
        ["Mês 12", 3672, 21, 3651, 43812, 17385, "Fim Free Tier", "Avaliar migração"],
        ["Mês 13", 3672, 1836, 1836, 45648, 100, "Infraestrutura híbrida", "Migração gradual"],
        ["Mês 14", 3672, 2754, 918, 46566, 50, "Mais recursos pagos", "Conforme receita"],
        ["Mês 15", 3672, 3672, 0, 46566, 0, "Infraestrutura completa", "Meta de $100k atingida"],
    ]
    
    # Preencher dados temporais
    for row, data in enumerate(temporal_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws_temporal.cell(row=row, column=col, value=value)
            cell.border = thin_border
            
            if col in [2, 3, 4, 5]:  # Valores monetários
                cell.alignment = currency_alignment
            elif col == 6:  # ROI
                cell.alignment = center_alignment
                if isinstance(value, (int, float)) and value > 1000:
                    cell.fill = savings_fill
                    cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.alignment = center_alignment if col in [1, 6] else Alignment(horizontal="left", vertical="center")
            
            # Destacar períodos importantes
            if data[0] in ["Mês 12", "Mês 15"]:
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    # Ajustar larguras temporais
    temporal_widths = [10, 15, 15, 15, 18, 12, 20, 18]
    for col, width in enumerate(temporal_widths, 1):
        ws_temporal.column_dimensions[get_column_letter(col)].width = width
    
    # ABA 3: RECOMENDAÇÕES
    ws_recommendations = wb.create_sheet("Recomendações")
    
    # Cabeçalho
    ws_recommendations['A1'] = "RECOMENDAÇÕES ESTRATÉGICAS"
    ws_recommendations['A1'].font = Font(bold=True, size=16)
    ws_recommendations.merge_cells('A1:D1')
    
    # Seção: Estratégia de Implementação
    ws_recommendations['A3'] = "ESTRATÉGIA DE IMPLEMENTAÇÃO"
    ws_recommendations['A3'].font = Font(bold=True, size=14)
    ws_recommendations['A3'].fill = header_fill
    ws_recommendations['A3'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_recommendations.merge_cells('A3:D3')
    
    recommendations_data = [
        ["Fase", "Período", "Infraestrutura", "Justificativa"],
        ["Validação", "Meses 1-6", "100% Free Tier", "Economia máxima durante desenvolvimento do MVP"],
        ["Crescimento", "Meses 7-12", "Free Tier + Serviços Pagos", "Adicionar recursos conforme necessidade"],
        ["Escala", "Meses 13+", "Infraestrutura Completa", "Migração quando receita justificar custos"],
    ]
    
    for row, data in enumerate(recommendations_data, 5):
        for col, value in enumerate(data, 1):
            cell = ws_recommendations.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if row == 5:  # Header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.alignment = center_alignment if col in [1, 2] else Alignment(horizontal="left", vertical="center")
    
    # Seção: Critérios de Migração
    ws_recommendations['A10'] = "CRITÉRIOS PARA MIGRAÇÃO"
    ws_recommendations['A10'].font = Font(bold=True, size=14)
    ws_recommendations['A10'].fill = header_fill
    ws_recommendations['A10'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_recommendations.merge_cells('A10:D10')
    
    criteria_data = [
        ["Métrica", "Limite Free Tier", "Quando Migrar", "Custo Adicional"],
        ["Clientes Ativos", "50-100", "> 100 clientes", "Conforme escala"],
        ["Receita Mensal", "Qualquer", "> $10.000", "$1.000-2.000/mês"],
        ["Execuções/Mês", "Ilimitado*", "> 100.000", "Recursos adicionais"],
        ["Uptime SLA", "Sem garantia", "99.9% necessário", "$500-1.000/mês"],
        ["Compliance", "Básico", "SOC2/ISO27001", "$2.000-3.000/mês"],
    ]
    
    for row, data in enumerate(criteria_data, 12):
        for col, value in enumerate(data, 1):
            cell = ws_recommendations.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if row == 12:  # Header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.alignment = center_alignment if col in [2, 4] else Alignment(horizontal="left", vertical="center")
    
    # Seção: Riscos e Mitigações
    ws_recommendations['A19'] = "RISCOS E MITIGAÇÕES"
    ws_recommendations['A19'].font = Font(bold=True, size=14)
    ws_recommendations['A19'].fill = header_fill
    ws_recommendations['A19'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_recommendations.merge_cells('A19:D19')
    
    risks_data = [
        ["Risco", "Probabilidade", "Impacto", "Mitigação"],
        ["Limite Free Tier", "Alta", "Médio", "Monitoramento contínuo de uso"],
        ["Performance", "Média", "Alto", "Otimização de código e queries"],
        ["Disponibilidade", "Baixa", "Alto", "Backup e procedimentos de recovery"],
        ["Escalabilidade", "Alta", "Alto", "Plano de migração preparado"],
        ["Suporte", "Média", "Médio", "Documentação detalhada"],
    ]
    
    for row, data in enumerate(risks_data, 21):
        for col, value in enumerate(data, 1):
            cell = ws_recommendations.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if row == 21:  # Header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.alignment = center_alignment
                # Colorir por nível de risco
                if col == 2:  # Probabilidade
                    if value == "Alta":
                        cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                    elif value == "Média":
                        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                elif col == 3:  # Impacto
                    if value == "Alto":
                        cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                    elif value == "Médio":
                        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    
    # Ajustar larguras das recomendações
    rec_widths = [15, 15, 20, 35]
    for col, width in enumerate(rec_widths, 1):
        ws_recommendations.column_dimensions[get_column_letter(col)].width = width
    
    # ABA 4: DASHBOARD EXECUTIVO
    ws_dashboard = wb.create_sheet("Dashboard Executivo")
    
    # Cabeçalho
    ws_dashboard['A1'] = "DASHBOARD EXECUTIVO - RESUMO FINANCEIRO"
    ws_dashboard['A1'].font = Font(bold=True, size=16)
    ws_dashboard.merge_cells('A1:F1')
    
    # KPIs Principais
    ws_dashboard['A3'] = "ECONOMIA TOTAL COM FREE TIER"
    ws_dashboard['A3'].font = Font(bold=True, size=14)
    ws_dashboard['A3'].fill = savings_fill
    ws_dashboard['A3'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_dashboard.merge_cells('A3:F3')
    
    kpis_data = [
        ["Métrica", "Valor", "Período", "Observação"],
        ["Economia Mensal", f"${total_savings:,}", "Meses 1-12", "Durante período free tier"],
        ["Economia Anual", f"${total_savings * 12:,}", "Primeiro ano", "Economia total no primeiro ano"],
        ["ROI Free Tier", f"{savings_percent:.0f}%", "Imediato", "Retorno sobre não-investimento"],
        ["Payback Period", "0 meses", "Imediato", "Sem investimento inicial"],
        ["Break-even", "$21/mês", "Custos mínimos", "Apenas domínio opcional"],
    ]
    
    for row, data in enumerate(kpis_data, 5):
        for col, value in enumerate(data, 1):
            cell = ws_dashboard.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if row == 5:  # Header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.alignment = center_alignment if col in [1, 2, 3] else Alignment(horizontal="left", vertical="center")
                if col == 2 and "$" in str(value):  # Valores monetários
                    cell.fill = savings_fill
                    cell.font = Font(color="FFFFFF", bold=True)
    
    # Conclusão
    ws_dashboard['A12'] = "CONCLUSÃO EXECUTIVA"
    ws_dashboard['A12'].font = Font(bold=True, size=14)
    ws_dashboard['A12'].fill = header_fill
    ws_dashboard['A12'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_dashboard.merge_cells('A12:F12')
    
    conclusion_text = [
        "• A versão Free Tier oferece economia de 99% nos custos de infraestrutura",
        "• Permite validação do modelo de negócio com investimento mínimo",
        "• Ideal para os primeiros 6-12 meses de operação",
        "• Migração gradual conforme crescimento da receita",
        "• ROI imediato devido à eliminação de custos fixos"
    ]
    
    for i, text in enumerate(conclusion_text, 14):
        cell = ws_dashboard.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_dashboard.merge_cells(f'A{i}:F{i}')
    
    # Ajustar larguras do dashboard
    dashboard_widths = [20, 15, 15, 30]
    for col, width in enumerate(dashboard_widths, 1):
        ws_dashboard.column_dimensions[get_column_letter(col)].width = width
    
    # Salvar arquivo
    filename = "/home/ubuntu/MoreFocus/Comparacao_Custos_Free_Tier.xlsx"
    wb.save(filename)
    print(f"Planilha de comparação criada: {filename}")
    
    # Resumo no console
    print(f"\n📊 RESUMO DA ANÁLISE:")
    print(f"💰 Custo Original: ${total_original:,}/mês")
    print(f"🆓 Custo Free Tier: ${total_free_tier}/mês")
    print(f"💵 Economia Mensal: ${total_savings:,}")
    print(f"📈 Economia Anual: ${total_savings * 12:,}")
    print(f"🎯 Economia Percentual: {savings_percent:.1f}%")
    
    return filename

if __name__ == "__main__":
    create_cost_comparison_spreadsheet()

