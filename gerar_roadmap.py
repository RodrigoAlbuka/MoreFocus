#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar

def create_roadmap_spreadsheet():
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Remover a planilha padrão
    wb.remove(wb.active)
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    subheader_fill = PatternFill(start_color="42A5F5", end_color="42A5F5", fill_type="solid")
    highlight_fill = PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid")
    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ABA 1: CRONOGRAMA GERAL
    ws_cronograma = wb.create_sheet("Cronograma Geral")
    
    # Cabeçalho
    ws_cronograma['A1'] = "ROADMAP MOREFOCUS - CRONOGRAMA DE IMPLEMENTAÇÃO"
    ws_cronograma['A1'].font = Font(bold=True, size=16)
    ws_cronograma.merge_cells('A1:R1')
    
    # Headers das colunas
    headers = ['Fase', 'Período', 'Atividade', 'Responsável', 'Status', 'Início', 'Fim', 
               'Duração (dias)', 'Dependências', 'Entregáveis', 'Riscos', 'Mitigação',
               'Orçamento (USD)', 'Receita Projetada', 'Clientes Meta', 'KPI Principal',
               'Meta KPI', 'Observações']
    
    for col, header in enumerate(headers, 1):
        cell = ws_cronograma.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Dados do cronograma
    cronograma_data = [
        # FASE 1: FUNDAÇÃO E MVP (Meses 1-3)
        ["Fase 1", "Mês 1", "Setup Infraestrutura AWS", "Fundador", "Planejado", "2025-01-01", "2025-01-14", 14, "-", "Infraestrutura básica", "Complexidade técnica", "Documentação detalhada", 5000, 0, 0, "Uptime", "95%", "Prioridade máxima"],
        ["Fase 1", "Mês 1", "Desenvolvimento APQ", "Fundador", "Planejado", "2025-01-15", "2025-01-31", 17, "Infraestrutura", "Agente de Prospecção", "Bugs iniciais", "Testes extensivos", 0, 0, 0, "Leads gerados", "50", "Primeiro agente"],
        ["Fase 1", "Mês 2", "Desenvolvimento ANE", "Fundador", "Planejado", "2025-02-01", "2025-02-14", 14, "APQ", "Agente de Nutrição", "Integração complexa", "APIs bem documentadas", 0, 0, 0, "Email open rate", "25%", "Automação de marketing"],
        ["Fase 1", "Mês 2", "Desenvolvimento AVC", "Fundador", "Planejado", "2025-02-15", "2025-02-28", 14, "ANE", "Agente de Vendas", "Lógica de negociação", "Cenários de teste", 0, 1000, 1, "Conversion rate", "5%", "Primeiro cliente"],
        ["Fase 1", "Mês 3", "Desenvolvimento AEI/ASM", "Fundador", "Planejado", "2025-03-01", "2025-03-15", 15, "AVC", "Agentes Entrega/Suporte", "Qualidade entrega", "Checklist qualidade", 0, 3000, 2, "Customer satisfaction", "85%", "Operação completa"],
        ["Fase 1", "Mês 3", "Desenvolvimento AFA", "Fundador", "Planejado", "2025-03-16", "2025-03-31", 16, "AEI/ASM", "Agente Financeiro", "Compliance fiscal", "Consultoria contábil", 2000, 8000, 5, "Margem líquida", "30%", "Automação financeira"],
        
        # FASE 2: VALIDAÇÃO E OTIMIZAÇÃO (Meses 4-6)
        ["Fase 2", "Mês 4", "Otimização Performance", "Fundador", "Planejado", "2025-04-01", "2025-04-15", 15, "Todos agentes", "Relatório performance", "Gargalos sistema", "Monitoramento contínuo", 1000, 12000, 8, "Response time", "<300ms", "Melhoria contínua"],
        ["Fase 2", "Mês 4-5", "Expansão Funcionalidades", "Fundador", "Planejado", "2025-04-16", "2025-05-15", 30, "Otimização", "Novas features", "Scope creep", "Roadmap rígido", 2000, 16000, 12, "Feature adoption", "80%", "Baseado em feedback"],
        ["Fase 2", "Mês 5-6", "Preparação Internacional", "Fundador", "Planejado", "2025-05-16", "2025-06-30", 46, "Expansão func.", "Documentação i18n", "Complexidade cultural", "Pesquisa mercado", 3000, 20000, 18, "Localization ready", "100%", "Foco Argentina/EUA"],
        
        # FASE 3: EXPANSÃO GEOGRÁFICA (Meses 7-12)
        ["Fase 3", "Mês 7-8", "Lançamento Argentina", "Fundador", "Planejado", "2025-07-01", "2025-08-31", 62, "Prep. Internacional", "Operação Argentina", "Instabilidade econômica", "Pricing flexível", 2000, 25000, 22, "Market penetration", "2%", "Mercado teste"],
        ["Fase 3", "Mês 9-10", "Lançamento EUA", "Fundador", "Planejado", "2025-09-01", "2025-10-31", 61, "Argentina", "Operação EUA", "Competição intensa", "Diferenciação clara", 5000, 40000, 36, "Enterprise clients", "5", "Mercado premium"],
        ["Fase 3", "Mês 11-12", "Preparação Europa", "Fundador", "Planejado", "2025-11-01", "2025-12-31", 61, "EUA", "GDPR compliance", "Regulamentações", "Consultoria legal", 8000, 70000, 58, "GDPR compliance", "100%", "Foco compliance"],
        
        # FASE 4: CONSOLIDAÇÃO E ESCALA (Meses 13-18)
        ["Fase 4", "Mês 13-14", "Lançamento Europa", "Fundador", "Planejado", "2026-01-01", "2026-02-28", 59, "Prep. Europa", "Operação Europa", "Fragmentação mercado", "Estratégia país-específica", 10000, 80000, 88, "EU market share", "1%", "Mercado regulado"],
        ["Fase 4", "Mês 15-16", "Otimização Receita", "Fundador", "Planejado", "2026-03-01", "2026-04-30", 61, "Europa", "Meta 100k/mês", "Saturação mercado", "Upsell/cross-sell", 5000, 100000, 126, "Monthly revenue", "100k USD", "Meta principal"],
        ["Fase 4", "Mês 17-18", "Preparação Futuro", "Fundador", "Planejado", "2026-05-01", "2026-06-30", 61, "Meta receita", "Roadmap 2027", "Complacência", "Inovação contínua", 3000, 120000, 172, "Growth rate", "15% MoM", "Sustentabilidade"],
    ]
    
    # Preencher dados
    for row, data in enumerate(cronograma_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws_cronograma.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = left_alignment if col in [3, 9, 10, 11, 12, 18] else center_alignment
            
            # Colorir por fase
            if data[0] == "Fase 1":
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            elif data[0] == "Fase 2":
                cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
            elif data[0] == "Fase 3":
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif data[0] == "Fase 4":
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    # Ajustar largura das colunas
    column_widths = [8, 10, 25, 12, 12, 12, 12, 12, 15, 20, 15, 15, 15, 15, 12, 15, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws_cronograma.column_dimensions[get_column_letter(col)].width = width
    
    # ABA 2: MÉTRICAS E KPIS
    ws_metricas = wb.create_sheet("Métricas e KPIs")
    
    # Cabeçalho
    ws_metricas['A1'] = "MÉTRICAS E KPIS - ACOMPANHAMENTO MENSAL"
    ws_metricas['A1'].font = Font(bold=True, size=16)
    ws_metricas.merge_cells('A1:P1')
    
    # Headers
    metricas_headers = ['Mês', 'Fase', 'Clientes Novos', 'Clientes Total', 'Receita Mensal (USD)', 
                       'MRR (USD)', 'Churn Rate (%)', 'CAC (USD)', 'LTV (USD)', 'NPS', 
                       'CSAT (%)', 'Uptime (%)', 'Lead Conversion (%)', 'Margem Líquida (%)', 
                       'Mercados Ativos', 'Status']
    
    for col, header in enumerate(metricas_headers, 1):
        cell = ws_metricas.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Dados das métricas (baseado na planilha CSV criada anteriormente)
    metricas_data = [
        [1, "Fundação", 1, 1, 1000, 800, 0, 1000, 5000, 30, 85, 95, 5, 30, 1, "Início"],
        [2, "Fundação", 1, 2, 3000, 2400, 0, 800, 5000, 35, 88, 98, 8, 40, 1, "Desenvolvimento"],
        [3, "Fundação", 3, 5, 8000, 6400, 0, 600, 5000, 40, 90, 99, 12, 50, 1, "MVP Completo"],
        [4, "Validação", 3, 8, 12000, 9600, 5, 550, 5200, 45, 91, 99.2, 15, 60, 1, "Otimização"],
        [5, "Validação", 4, 12, 16000, 12800, 3, 500, 5400, 50, 92, 99.5, 18, 65, 1, "Refinamento"],
        [6, "Validação", 6, 18, 20000, 16000, 2, 450, 5600, 55, 93, 99.7, 20, 70, 1, "Preparação Int'l"],
        [7, "Expansão", 4, 22, 25000, 20000, 3, 500, 5800, 50, 91, 99.5, 18, 68, 2, "Argentina"],
        [8, "Expansão", 6, 28, 30000, 24000, 4, 520, 6000, 52, 92, 99.6, 19, 70, 2, "Consolidação ARG"],
        [9, "Expansão", 8, 36, 40000, 32000, 3, 480, 6200, 55, 93, 99.7, 22, 72, 3, "EUA Lançamento"],
        [10, "Expansão", 10, 46, 50000, 40000, 2, 450, 6400, 58, 94, 99.8, 25, 74, 3, "EUA Crescimento"],
        [11, "Expansão", 12, 58, 60000, 48000, 2, 420, 6600, 60, 95, 99.8, 28, 75, 3, "Prep. Europa"],
        [12, "Expansão", 14, 72, 70000, 56000, 3, 400, 6800, 62, 95, 99.9, 30, 76, 3, "Consolidação"],
        [13, "Consolidação", 16, 88, 80000, 64000, 2, 380, 7000, 65, 96, 99.9, 32, 77, 4, "Europa Launch"],
        [14, "Consolidação", 18, 106, 90000, 72000, 2, 360, 7200, 67, 96, 99.9, 34, 78, 4, "Europa Growth"],
        [15, "Consolidação", 20, 126, 100000, 80000, 2, 350, 7400, 70, 97, 99.9, 35, 79, 4, "Meta Atingida"],
        [16, "Escala", 22, 148, 110000, 88000, 2, 340, 7600, 72, 97, 99.9, 36, 80, 4, "Otimização"],
        [17, "Escala", 24, 172, 120000, 96000, 2, 330, 7800, 74, 98, 99.9, 37, 81, 4, "Crescimento"],
        [18, "Escala", 26, 198, 130000, 104000, 2, 320, 8000, 75, 98, 99.9, 38, 82, 4, "Sustentável"],
    ]
    
    # Preencher dados das métricas
    for row, data in enumerate(metricas_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws_metricas.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment
            
            # Destacar marcos importantes
            if data[15] in ["MVP Completo", "Meta Atingida", "Sustentável"]:
                cell.fill = highlight_fill
                if col == 16:  # Status column
                    cell.font = Font(bold=True, color="FFFFFF")
    
    # Ajustar largura das colunas
    metricas_widths = [6, 12, 12, 12, 15, 12, 12, 10, 10, 8, 10, 10, 15, 15, 12, 15]
    for col, width in enumerate(metricas_widths, 1):
        ws_metricas.column_dimensions[get_column_letter(col)].width = width
    
    # ABA 3: ORÇAMENTO E INVESTIMENTOS
    ws_orcamento = wb.create_sheet("Orçamento")
    
    # Cabeçalho
    ws_orcamento['A1'] = "ORÇAMENTO E INVESTIMENTOS - BREAKDOWN DETALHADO"
    ws_orcamento['A1'].font = Font(bold=True, size=16)
    ws_orcamento.merge_cells('A1:H1')
    
    # Seção: Investimento Inicial
    ws_orcamento['A3'] = "INVESTIMENTO INICIAL"
    ws_orcamento['A3'].font = Font(bold=True, size=14)
    ws_orcamento['A3'].fill = subheader_fill
    ws_orcamento.merge_cells('A3:H3')
    
    investimento_headers = ['Item', 'Descrição', 'Quantidade', 'Valor Unitário (USD)', 'Total (USD)', 'Categoria', 'Prioridade', 'Observações']
    for col, header in enumerate(investimento_headers, 1):
        cell = ws_orcamento.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    investimento_data = [
        ["Infraestrutura AWS", "Setup inicial multi-região", 1, 10000, 10000, "Infraestrutura", "Alta", "Inclui 3 meses"],
        ["Desenvolvimento n8n", "Licenças e setup", 1, 5000, 5000, "Software", "Alta", "Self-hosted premium"],
        ["Ferramentas DevOps", "CI/CD, monitoramento", 1, 3000, 3000, "Ferramentas", "Média", "Prometheus, Grafana"],
        ["Marketing Inicial", "Branding, website", 1, 8000, 8000, "Marketing", "Alta", "Identidade visual"],
        ["Legal e Compliance", "Contratos, GDPR", 1, 4000, 4000, "Legal", "Alta", "Consultoria jurídica"],
        ["Reserva Emergência", "Capital de giro", 1, 20000, 20000, "Reserva", "Alta", "4 meses operação"],
    ]
    
    total_investimento = 0
    for row, data in enumerate(investimento_data, 5):
        for col, value in enumerate(data, 1):
            cell = ws_orcamento.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment if col in [3, 4, 5] else left_alignment
            if col == 5:  # Total column
                total_investimento += value
    
    # Total do investimento
    ws_orcamento.cell(row=11, column=4, value="TOTAL INVESTIMENTO:").font = Font(bold=True)
    ws_orcamento.cell(row=11, column=5, value=total_investimento).font = Font(bold=True)
    ws_orcamento.cell(row=11, column=5).fill = highlight_fill
    
    # Seção: Custos Operacionais Mensais
    ws_orcamento['A13'] = "CUSTOS OPERACIONAIS MENSAIS"
    ws_orcamento['A13'].font = Font(bold=True, size=14)
    ws_orcamento['A13'].fill = subheader_fill
    ws_orcamento.merge_cells('A13:H13')
    
    # Headers para custos operacionais
    for col, header in enumerate(investimento_headers, 1):
        cell = ws_orcamento.cell(row=14, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    custos_data = [
        ["AWS Infrastructure", "Compute, storage, network", 1, 2000, 2000, "Infraestrutura", "Alta", "Escala com uso"],
        ["n8n Hosting", "Self-hosted premium", 1, 500, 500, "Software", "Alta", "Inclui suporte"],
        ["Marketing Digital", "Ads, content, tools", 1, 3000, 3000, "Marketing", "Alta", "Multi-mercado"],
        ["Ferramentas SaaS", "CRM, analytics, etc", 1, 800, 800, "Ferramentas", "Média", "Stack completo"],
        ["Compliance", "Auditorias, certificações", 1, 1000, 1000, "Legal", "Alta", "GDPR, SOC2"],
        ["Contingência", "Imprevistos", 1, 700, 700, "Reserva", "Média", "10% dos custos"],
    ]
    
    total_mensal = 0
    for row, data in enumerate(custos_data, 15):
        for col, value in enumerate(data, 1):
            cell = ws_orcamento.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment if col in [3, 4, 5] else left_alignment
            if col == 5:  # Total column
                total_mensal += value
    
    # Total mensal
    ws_orcamento.cell(row=21, column=4, value="TOTAL MENSAL:").font = Font(bold=True)
    ws_orcamento.cell(row=21, column=5, value=total_mensal).font = Font(bold=True)
    ws_orcamento.cell(row=21, column=5).fill = highlight_fill
    
    # Ajustar larguras
    orcamento_widths = [20, 25, 12, 18, 15, 15, 12, 20]
    for col, width in enumerate(orcamento_widths, 1):
        ws_orcamento.column_dimensions[get_column_letter(col)].width = width
    
    # ABA 4: RECURSOS E RESPONSABILIDADES
    ws_recursos = wb.create_sheet("Recursos")
    
    # Cabeçalho
    ws_recursos['A1'] = "RECURSOS E RESPONSABILIDADES"
    ws_recursos['A1'].font = Font(bold=True, size=16)
    ws_recursos.merge_cells('A1:J1')
    
    # Headers
    recursos_headers = ['Agente/Sistema', 'Função Principal', 'Responsabilidades', 'Tecnologias', 
                       'Dependências', 'SLA', 'Monitoramento', 'Backup/Recovery', 'Escalabilidade', 'Observações']
    
    for col, header in enumerate(recursos_headers, 1):
        cell = ws_recursos.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    recursos_data = [
        ["APQ - Agente Prospecção", "Lead Generation", "Scraping, qualificação, scoring", "n8n, Python, APIs", "Infraestrutura", "99.5%", "24/7", "Diário", "Auto-scaling", "Primeiro agente"],
        ["ANE - Agente Nutrição", "Email Marketing", "Sequências, segmentação, personalização", "n8n, Mailgun, CRM", "APQ", "99.5%", "24/7", "Diário", "Auto-scaling", "Integração CRM"],
        ["AVC - Agente Vendas", "Sales Automation", "Demos, propostas, negociação", "n8n, CRM, DocuSign", "ANE", "99.9%", "24/7", "Diário", "Manual", "Lógica complexa"],
        ["AEI - Agente Entrega", "Implementation", "Auditoria, desenvolvimento, deploy", "n8n, Git, AWS", "AVC", "99.9%", "Business hours", "Diário", "Manual", "Qualidade crítica"],
        ["ASM - Agente Suporte", "Support & Monitoring", "Tickets, monitoramento, alertas", "n8n, Zendesk, APM", "Todos", "99.9%", "24/7", "Tempo real", "Auto-scaling", "SLA crítico"],
        ["AFA - Agente Financeiro", "Finance & Admin", "Faturamento, cobrança, relatórios", "n8n, ERP, Banking APIs", "Todos", "99.9%", "Business hours", "Diário", "Manual", "Compliance"],
        ["Infraestrutura AWS", "Cloud Platform", "Compute, storage, network, security", "EKS, RDS, S3, VPC", "-", "99.9%", "24/7", "Contínuo", "Auto-scaling", "Multi-região"],
        ["n8n Platform", "Workflow Engine", "Orquestração, execução, monitoramento", "Node.js, PostgreSQL", "AWS", "99.9%", "24/7", "Diário", "Horizontal", "Core system"],
        ["Monitoramento", "Observability", "Métricas, logs, alertas, dashboards", "Prometheus, Grafana, ELK", "Infraestrutura", "99.5%", "24/7", "Tempo real", "Auto-scaling", "Visibilidade total"],
        ["Segurança", "Security & Compliance", "Auth, encryption, audit, compliance", "IAM, KMS, CloudTrail", "Todos", "100%", "24/7", "Contínuo", "Automático", "Zero trust"],
    ]
    
    # Preencher dados
    for row, data in enumerate(recursos_data, 4):
        for col, value in enumerate(data, 1):
            cell = ws_recursos.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Destacar agentes de IA
            if "Agente" in data[0]:
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    # Ajustar larguras
    recursos_widths = [25, 18, 30, 20, 15, 8, 15, 15, 15, 20]
    for col, width in enumerate(recursos_widths, 1):
        ws_recursos.column_dimensions[get_column_letter(col)].width = width
    
    # ABA 5: DASHBOARD EXECUTIVO
    ws_dashboard = wb.create_sheet("Dashboard Executivo")
    
    # Cabeçalho
    ws_dashboard['A1'] = "DASHBOARD EXECUTIVO - VISÃO CONSOLIDADA"
    ws_dashboard['A1'].font = Font(bold=True, size=16)
    ws_dashboard.merge_cells('A1:F1')
    
    # KPIs Principais
    ws_dashboard['A3'] = "KPIS PRINCIPAIS"
    ws_dashboard['A3'].font = Font(bold=True, size=14)
    ws_dashboard['A3'].fill = subheader_fill
    
    kpis_principais = [
        ["Meta Receita Mensal", "USD 100.000", "Mês 15"],
        ["ROI 18 meses", "1.370%", "Projetado"],
        ["Payback Period", "4 meses", "Break-even"],
        ["Margem Líquida Target", ">70%", "Sustentável"],
        ["Clientes Target Mês 18", "200", "Global"],
        ["Mercados Ativos", "4", "BR, AR, US, EU"],
    ]
    
    for row, (kpi, valor, status) in enumerate(kpis_principais, 5):
        ws_dashboard.cell(row=row, column=1, value=kpi).font = Font(bold=True)
        ws_dashboard.cell(row=row, column=2, value=valor).font = Font(bold=True, color="FF5722")
        ws_dashboard.cell(row=row, column=3, value=status)
    
    # Marcos Críticos
    ws_dashboard['A12'] = "MARCOS CRÍTICOS"
    ws_dashboard['A12'].font = Font(bold=True, size=14)
    ws_dashboard['A12'].fill = subheader_fill
    
    marcos = [
        ["Mês 3", "MVP Completo", "5 clientes, todos agentes funcionais"],
        ["Mês 6", "Validação", "18 clientes, processos otimizados"],
        ["Mês 9", "EUA Launch", "36 clientes, 3 mercados ativos"],
        ["Mês 12", "Consolidação", "72 clientes, preparação Europa"],
        ["Mês 15", "Meta Atingida", "126 clientes, USD 100k/mês"],
        ["Mês 18", "Sustentável", "200 clientes, crescimento estável"],
    ]
    
    for row, (mes, marco, descricao) in enumerate(marcos, 14):
        ws_dashboard.cell(row=row, column=1, value=mes).font = Font(bold=True)
        ws_dashboard.cell(row=row, column=2, value=marco).font = Font(bold=True, color="1A237E")
        ws_dashboard.cell(row=row, column=3, value=descricao)
    
    # Riscos e Mitigações
    ws_dashboard['A21'] = "PRINCIPAIS RISCOS E MITIGAÇÕES"
    ws_dashboard['A21'].font = Font(bold=True, size=14)
    ws_dashboard['A21'].fill = subheader_fill
    
    riscos = [
        ["Técnico", "Falhas de infraestrutura", "Redundância multi-região, monitoramento 24/7"],
        ["Mercado", "Competição intensa", "Diferenciação por automação total"],
        ["Operacional", "Dependência fundador", "Documentação extensiva, automação máxima"],
        ["Financeiro", "Fluxo de caixa", "Reserva de emergência, cobrança automatizada"],
        ["Regulatório", "Mudanças compliance", "Monitoramento contínuo, adaptabilidade"],
    ]
    
    for row, (tipo, risco, mitigacao) in enumerate(riscos, 23):
        ws_dashboard.cell(row=row, column=1, value=tipo).font = Font(bold=True, color="FF5722")
        ws_dashboard.cell(row=row, column=2, value=risco)
        ws_dashboard.cell(row=row, column=3, value=mitigacao)
    
    # Ajustar larguras do dashboard
    dashboard_widths = [15, 25, 40]
    for col, width in enumerate(dashboard_widths, 1):
        ws_dashboard.column_dimensions[get_column_letter(col)].width = width
    
    # Salvar arquivo
    filename = "/home/ubuntu/MoreFocus/MoreFocus_Roadmap_Detalhado.xlsx"
    wb.save(filename)
    print(f"Planilha criada com sucesso: {filename}")
    
    return filename

if __name__ == "__main__":
    create_roadmap_spreadsheet()

